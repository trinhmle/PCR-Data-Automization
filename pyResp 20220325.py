#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Fri Mar 25 03:50:14 2022

@author: trinhle
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Mar 23 17:14:56 2022

@author: trinhle
"""

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Mar 22 17:24:23 2022

@author: trinhle
"""
#pip install openpyxl

##import openpyxl
##from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
##from openpyxl.styles import Font
import timeit

start = timeit.default_timer()

fileRT = './Resp_20220325_01_TL RT.xlsx'
fileSampleLog = './Resp_20220325_01_TL.xlsx'
fileQuantSummary = './Resp_20220325_01_TL -  Quantification Summary.xlsx'

wbRT = load_workbook(fileRT)
wbSL = load_workbook(fileSampleLog)
##wbQS = load_workbook(fileQuantSummary)

wsRT = wbRT['Sheet2']
#wsRTpositive = wbRT.create_sheet('RESULTS')
wsRTpositive = wbRT['RESULTS']
#wsRTresults = wbRT.create_sheet('Log')
wsRTresults = wbRT['Log']



for row in wsRTpositive['A1:P500']:
    for cell in row:
        cell.value = None
 
for row in wsRTresults['A1:P500']:
    for cell in row:
        cell.value = None
        
wsSLsample = wbSL['Sample Log Value']


##!!!!!!!!PLEASE CHANGE THE SHEET NAMES WHEN YOU HAVE THE REAL FILES!!!!!!!!!



colRE = 1
charRE = get_column_letter(colRE)

colType = 2
charType = get_column_letter(colType)

colName = 3
charName = get_column_letter(colName)

colReview = 9
charReview = get_column_letter(colReview)
print(charReview)

colCT = 4
charCT = get_column_letter(colCT)

colPositive = 7
charPositive = get_column_letter(colPositive)


posRE = []
posName = []
posCT = []

print('------POSITIVE RESULTS-----')
## grab all positve results
##-------------------------------POSTIVE-------------------------------
for row in range (1, 500):
    
#!!!!change range for more samples
    
    
    cellNumberPositive = charPositive + str(row)
    cellNumberRE = charRE + str(row)
    cellNumberName = charName + str(row)
    cellNumberType = charType + str(row)
    cellNumberA = str('A') + str(row)
    cellNumberCT = charCT + str(row)
    ##print(ws[cellNumber].value)
    ##print(cellNumber)
    
    if wsRT[cellNumberType].value == 'Unkn':
        
       
        
        if wsRT[cellNumberPositive].value == 'POSITIVE':
            
            
            if wsRT[cellNumberRE].value != 'RE3-IC (EAV)':
                    print('%-15s' % str(wsRT[cellNumberName].value) + ' ' + '%-20s' % str(wsRT[cellNumberRE].value) + ' ' + '%-5s' % str("{:.2f}".format(wsRT[cellNumberCT].value)))
                    posRE.append(str(wsRT[cellNumberRE].value))
                    posName.append(str(wsRT[cellNumberName].value))
                    posCT.append(str("{:.2f}".format(wsRT[cellNumberCT].value)))

## APPEND DATA TO WORKBOOK  
##-------------------------------APPEND POSITIVE-------------------------------


cellNumberPositive = len(posName)

wsRTpositive.cell(1, 1, 'Positive Results')

for cellNumber in range(cellNumberPositive):
    wsRTpositive.cell(row = cellNumber + 2, column = 1, value = posName[cellNumber])
    
for cellNumber in range(cellNumberPositive):
    wsRTpositive.cell(row = cellNumber + 2, column = 2, value = posRE[cellNumber])
    
for cellNumber in range(cellNumberPositive):
    wsRTpositive.cell(row = cellNumber + 2, column = 3, value = posCT[cellNumber])
    
#print()

##-------------------------------REVIEW-------------------------------

reviewName = []
reviewRE = []
reviewCT = []

print('-----REIVEW RESULTS-----')                
##check review results

for row in range (1, 500):   
##change range for more samples
    cellNumberType = charType + str(row)
    cellNumberName = charName + str(row)
    cellNumberReview = charReview + str(row)
    cellNumberCT = charCT + str(row)
    cellNumberRE = charRE +str(row)
    
    if wsRT[cellNumberType].value == 'Unkn':
        if wsRT[cellNumberReview].value == 'Review':
            print('%-15s' % str(wsRT[cellNumberName].value) + ' ' + '%-20s' % str(wsRT[cellNumberRE].value) + ' ' + '%-5s' % str("{:.2f}".format(wsRT[cellNumberCT].value)))
            reviewName.append(str(wsRT[cellNumberName].value))
            reviewRE.append(str(wsRT[cellNumberRE].value))
            reviewCT.append(str("{:.2f}".format(wsRT[cellNumberCT].value)))
            

## APPEND DATA TO WORKBOOK
##-------------------------------APPPEND REVIEW-------------------------------

##wsRTreview = wbRT.create_sheet('REVIEW RESULTS')

cellNumberReview = len(reviewName)

wsRTpositive.cell(1, 5, 'Review')


for cellNumber in range(cellNumberReview):
    wsRTpositive.cell(row = cellNumber + 2, column = 5, value = reviewName[cellNumber])

for cellNumber in range(cellNumberReview):
    wsRTpositive.cell(row = cellNumber + 2, column = 6, value = reviewRE[cellNumber])
    
for cellNumber in range(cellNumberReview):
    wsRTpositive.cell(row = cellNumber + 2, column = 7, value = reviewCT[cellNumber])
    

    

#print()

##-------------------------------CT BUT NEGATIVE-------------------------------
negRE = []
negName = []
negCT = []

print('-----CT but Negative-----')
for row in range (1, 500):   
##change range for more samples
    cellNumberType = charType + str(row)
    cellNumberName = charName + str(row)
    cellNumberCT = charCT + str(row)
    cellNumberRE = charRE + str(row)
    cellNumberPositive = charPositive + str(row)
    
    
    ##if wsRT[cellNumberPositive].value == 'NEGATIVE':
        
    if wsRT[cellNumberType].value == 'Unkn':
        if wsRT[cellNumberCT].value != None :
            #print(str(wsRT[cellNumberName].value + ' ' + str(wsRT[cellNumberCT].value)) + ' ' + str(wsRT[cellNumberPositive].value))
            if wsRT[cellNumberPositive].value == 'NEGATIVE':
         
                if wsRT[cellNumberRE].value != 'RE3-IC (EAV)':
                    
                    print('%-15s' % str(wsRT[cellNumberName].value) + ' ' + '%-20s' % str(wsRT[cellNumberRE].value) + '%-5s' % str("{:.2f}".format(wsRT[cellNumberCT].value)))
                    negName.append(str(wsRT[cellNumberName].value))
                    negRE.append(str(wsRT[cellNumberRE].value))
                    negCT.append(str("{:.2f}".format(wsRT[cellNumberCT].value)))


##-------------------------------APPEND CT BUT NEGATIVE-------------------------------

cellNumberNegative = len(negName)



wsRTpositive.cell(1, 9, 'CT but Negative')

for cellNumber in range(cellNumberNegative):
    wsRTpositive.cell(row = cellNumber + 2, column = 9, value = negName[cellNumber])

for cellNumber in range(cellNumberNegative):
    wsRTpositive.cell(row = cellNumber + 2, column = 10, value = negRE[cellNumber])

for cellNumber in range(cellNumberNegative):
    wsRTpositive.cell(row = cellNumber + 2, column = 11, value = negCT[cellNumber])

#print()


## try to append new data to sample log workbook

wbRT.save(fileRT)


##-------------APPEND TO SAMPLE LOG--------------

charE = 'E'


##CLEAR SAMPLE LOG

for row in range(6, 51):
    cellNumber = charE + str(row)
    wsSLsample[cellNumber].value = None




##FIGURING OUT STUFF AND STUFF


name = []
for row in range (2, 11):
    cellNumber = 'A' + str(row)
    i = len(name)
    
    check = 0
    
    for size in range(i):
        
        if wsRTpositive[cellNumber].value == name[size]:
            
            check = 1
        
    if check == 0:
        
        name.append(wsRTpositive[cellNumber].value)


for i in range(len(name)):
    
    a = []
  
    
    for row in range(2,11):
        
        cellNumberA = 'A' + str(row)
        cellNumberB = 'B' + str(row)
        
        if wsRTpositive[cellNumberA].value == name[i]:
            
                
            a.append(str(wsRTpositive[cellNumberB].value))
            
    
    wsRTresults.cell(row = i + 1, column = 1, value = name[i])   
    
    
    b = None
            
    for j in range(len(a)):
                
        b = str(b) + ', ' + a[j]
    
    
    
    for k in range(1, 9):
        r = 'RE' + str(k) + '-'
        b = b.replace(str(r), '')
        
    b = b.replace('None, ', '')
       
    wsRTresults.cell(row = i + 1, column = 2, value = b)
            
            
        
wbRT.save(fileRT)        
        
            
            
for rowRT in range(len(name)):
    
    cellNumberA = 'A' + str((rowRT)+1)
    cellNumberB = 'B' + str((rowRT)+1)
    
    for rowSL in range(6, 50):
        
        
        cellNumberC = 'C' + str(rowSL)
        cellNumberE = 'E' + str(rowSL)
    
        
        if str(wsRTresults[cellNumberA].value) == str(wsSLsample[cellNumberC].value):
            
            wsSLsample[cellNumberE] = str(wsRTresults[cellNumberB].value)
    

##-----------IF BLANK THEN NEGATIVE----------
for row in range(6, 51):
    
    cellNumber = charE + str(row)
    
    if wsSLsample[cellNumber].value == None:
        
        wsSLsample[cellNumber].value = 'Neg'
    

wbSL.save(fileSampleLog)


stop = timeit.default_timer()

print('Run Time: ', stop - start) 





